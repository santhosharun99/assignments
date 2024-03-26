#NAME : SANTHOSH ARUNAGIRI
#ID : 201586816

from openpyxl import load_workbook

class InvalidAgent(Exception):
    """Exception raised when an invalid agent number is provided.
    Attributes:
        message -- explanation of the error
    """
    pass

class InvalidScoreVectorLength(Exception):
    """Exception raised when the length of the score vector is invalid.
    Attributes:
        message -- explanation of the error
    """
    pass

class InvalidTieKey(Exception):
    """Exception raised when an invalid tie-breaking key is provided.
    Attributes:
        message -- explanation of the error
    """
    pass

def generatePreferences(values):
    """Generate a preference profile from a worksheet of values.
    Parameters:
    values (worksheet): a worksheet containing the values for the alternatives
    Returns:
    dict: a dictionary of preferences, where the keys are agent numbers and the values are lists of preferences
    """
    preferences = {}
    for i in range(1, values.max_row + 1):
        agent_preferences = []
        for j in range(1, values.max_column + 1):
            # Check if the cell exists
            if values.max_row < i or values.max_column < j:
                continue
            # Add the alternative to the list of preferences for the agent
            agent_preferences.append(j)
        # Sort the list of preferences based on the values in the worksheet
        agent_preferences.sort(key=lambda x: (values[i][x - 1].value, x), reverse=True)
        preferences[i] = agent_preferences
    return preferences

def scoreCalculator(preferences, scoreVector, tieBreak):
    """Calculate the scores for the alternatives based on the given preferences and score vector.
        Parameters:
        preferences (dict): a dictionary of preferences, where the keys are agent numbers and the values are lists of preferences
        scoreVector (list): a list of scores to be assigned to the alternatives
        tieBreak (int, str): an integer representing the agent number to use for tie-breaking, or a string representing the tie-breaking rule to use (either "max" or "min")
        Returns:
        int: the alternative with the highest score
        """
    score_dic = {}
    for each_preference in preferences.values():
        for i in range(len(each_preference)):
            if each_preference[i] in score_dic.keys():
                score_dic[each_preference[i]] = scoreVector[i] + score_dic[each_preference[i]]
            else:
                score_dic[each_preference[i]] = scoreVector[i]
    return tieBreaker(preferences, score_dic, tieBreak)

def tieBreaker(preferences, score_dic, tieBreak):
    """Break ties between alternatives with the same score.
    Parameters:
    preferences (dict): a dictionary of preferences, where the keys are agent numbers and the values are lists of preferences
    score_dic (dict): a dictionary of scores, where the keys are alternative numbers and the values are scores
    tieBreak (int, str): an integer representing the agent number to use for tie-breaking, or a string representing the tie-breaking rule to use (either "max" or "min")
    Returns:
    int: the alternative with the highest score after tie-breaking
    """
    try:
        same_count = []
        for keys in score_dic.keys():
            if score_dic[keys] == max(score_dic.values()):
                same_count.append(keys)
        if list(score_dic.values()).count(max(score_dic.values())) > 1:
            if isinstance(tieBreak, int):
                if tieBreak not in preferences.keys():
                    raise InvalidTieKey
                else:
                    for i in preferences[tieBreak]:
                        if i in same_count:
                            return i
            elif tieBreak == "max":
                return max(same_count)
            elif tieBreak == "min":
                return min(same_count)
            else:
                raise InvalidTieKey
        else:
            return same_count[0]
    except InvalidTieKey:
        print("Incorrect Tie Break value")
        

def dictatorship(preferences, agent):
        """Determine the winner using the dictatorship voting rule.
        Parameters:
        preferences (dict): a dictionary of preferences, where the keys are agent numbers and the values are lists of preferences
        agent (int): the number of the agent whose preference will be used to determine the winner
        Returns:
        int: the winning alternative
        """
        return preferences[agent][0]
    
def scoringRule(preferences, scoreVector, tieBreak):
    """Determine the winner using the scoring rule voting method.
        Parameters:
        preferences (dict): a dictionary of preferences, where the keys are agent numbers and the values are lists of preferences
        scoreVector (list): a list of scores to be assigned to the alternatives
        tieBreak (int, str): an integer representing the agent number to use for tie-breaking, or a string representing the tie-breaking rule to use (either "max" or "min")
        Returns:
        int: the winning alternative
        """
    try:
        if len(scoreVector) != len(preferences[1]):
            raise InvalidScoreVectorLength
        else:
            scoreVector.sort()
            scoreVector = scoreVector[::-1]
#             print("s",scoreVector)
            return scoreCalculator(preferences, scoreVector, tieBreak)
    except InvalidScoreVectorLength:
        print("Incorrect input")
        return False
        
def plurality(preferences, tieBreak):
    """Determine the winner using the plurality voting method.
    Parameters:
    preferences (dict): a dictionary of preferences, where the keys are agent numbers and the values are lists of preferences
    tieBreak (int, str): an integer representing the agent number to use for tie-breaking, or a string representing the tie-breaking rule to use (either "max" or "min")
    Returns:
    int: the winning alternative
    """
    scoreVector = [1]+[0 for i in range(len(list(preferences.values())[0])-1)]
    return scoreCalculator(preferences, scoreVector, tieBreak)
  
def veto(preferences, tieBreak):
    """Determine the winner using the veto voting method.
    Parameters:
    preferences (dict): a dictionary of preferences, where the keys are agent numbers and the values are lists of preferences
    tieBreak (int, str): an integer representing the agent number to use for tie-breaking, or a string representing the tie-breaking rule to use (either "max" or "min")
    Returns:
    int: the winning alternative
    """
    scoreVector = [1 for i in range(len(list(preferences.values())[0])-1)]+[0]
    return scoreCalculator(preferences, scoreVector, tieBreak)
    
def borda(preferences, tieBreak):
    """Determine the winner using the Borda voting method.
    Parameters:
    preferences (dict): a dictionary of preferences, where the keys are agent numbers and the values are lists of preferences
    tieBreak (int, str): an integer representing the agent number to use for tie-breaking, or a string representing the tie-breaking rule to use (either "max" or "min")
    Returns:
    int: the winning alternative
    """
    scoreVector = list(range(len(list(preferences.values())[0])-1,-1,-1))
    return scoreCalculator(preferences, scoreVector, tieBreak)

def harmonic(preferences, tieBreak):
    """Determine the winner using the harmonic voting method.
    Parameters:
    preferences (dict): a dictionary of preferences, where the keys are
    agent numbers and the values are lists of preferences
    tieBreak (int, str): an integer representing the agent number to use for tie-breaking, or a string representing the tie-breaking rule to use (either "max" or "min")
    Returns:
    int: the winning alternative
    """
    scoreVector = [1/i for i in range(1,len(list(preferences.values())[0])+1)]
    return scoreCalculator(preferences, scoreVector, tieBreak)

def STV(preferences, tieBreak):
    """Determine the winner using the Single Transferable Vote (STV) method.
    Parameters:
    preferences (dict): a dictionary of preferences, where the keys are agent numbers and the values are lists of preferences
    tieBreak (int, str): an integer representing the agent number to use for tie-breaking, or a string representing the tie-breaking rule to use (either "max" or "min")
    Returns:
    int: the winning alternative
    """
    scoreVector = [1]+[0 for i in range(len(list(preferences.values())[0])-1)]
    return scoreCalculator(preferences, scoreVector, tieBreak)

def rangeVoting(values, tieBreak):
    """Determine the winner using the Range Voting method.
        Parameters:
        preferences (dict): a dictionary of preferences, where the keys are agent numbers and the values are lists of preferences
        tieBreak (int, str): an integer representing the agent number to use for tie-breaking, or a string representing the tie-breaking rule to use (either "max" or "min")
        Returns:
        int: the winning alternative
        Raises:
        InvalidAgent: if an agent number is not found in the preferences dictionary
        InvalidTieKey: if the tieBreak value is not a valid agent number or tie-breaking rule
        """
#     lis = [i.value for row in values.rows for i in row]
#     print(list)
    lis = []
    d = {}
    for row in values.rows:
        valuation = [data.value for data in row]
        lis.append(valuation)
    for i in range(len(lis[0])):
        s = 0
        for j in range(len(lis)):
            s += lis[j][i]
        d.update({i+1:s})
    return tieBreaker(generatePreferences(values), d, tieBreak)